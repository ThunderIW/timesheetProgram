import polars as pl
from openpyxl import *


import databaseManagement as db

def create_CSV_And_Excel():
    # 1. If retrieve_full_report() returns a list of dicts:
    report = db.retrieve_full_report()
    print(report)

    columns = [
        "Project Code",
        "product Name",
        "emp Code",
        "first Name",
        "last Name",
        "rate Per Hour",
        "total Hours  Worked",
        "project Budget",
        "total Cost"
    ]

    # Build DataFrame
    df = pl.DataFrame(report,schema=columns,orient="row")
    print(df)
    df.write_excel("timesheet_report.xlsx")
    #df.write_csv("timesheet_report.csv")
    print("Excel file saved as timesheet_report.xlsx")





def modify_excel_file(excel_file:str):

    wb=load_workbook(excel_file)
    print(wb.sheetnames)
    if "Sheet" in wb.sheetnames:
        ws1 = wb['Sheet1']
        ws1.title="DATA"
    else:
        ws1=wb['DATA']

    last_row = ws1.max_row

    if "ProjectInfo" not in wb.sheetnames:
        ws2=wb.create_sheet(index=1,title="ProjectInfo")
        #last_row=ws1.max_row
    else:
        ws2=wb["ProjectInfo"]
    #this is for the worksheet

    ws1[f'I{last_row+1}'] = f"=SUM(I2:I{last_row})"
    unique_project_codes=list(set(cell.value for cell in ws1['A'] if cell.value is not None))
    print(unique_project_codes)


    headers=["Project code","total cost"," Project Budget","Remaining  budget"]
    for col_num, header in enumerate(headers,1):
        ws2.cell(row=1, column=col_num, value=header)
    wb.save(excel_file)



#create_CSV_And_Excel()

print(0.02<0)



